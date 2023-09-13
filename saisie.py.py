from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import load_workbook

# Configuration de Selenium
def configure_browser():
    options = webdriver.ChromeOptions()
    return webdriver.Chrome(options=options)

# Accepter les cookies
def accept_cookies(driver, url):
    driver.get(url)
    wait = WebDriverWait(driver, 45)
    time.sleep(10)
    accept_all_button = wait.until(EC.visibility_of_element_located((By.ID, "axeptio_btn_acceptAll")))
    accept_all_button.click()

# Remplir le formulaire
def fill_form(driver, quantity_value, prenom_value, nom_value, texte_value, email_value):
    # Sélectionner l'élément input par son ID
    quantity_input = driver.find_element(By.ID, "quantity-selector-id_9363320")

    # Effacer la valeur actuelle (0) et saisir la nouvelle quantité depuis le fichier Excel
    quantity_input.clear()
    quantity_input.send_keys(str(quantity_value))

    # Cliquez sur le bouton "Étape suivante" pour passer à la page suivante du formulaire
    next_button = driver.find_element(By.XPATH, '//button[@data-ux="Forms_EVENT_NextStep"]')
    next_button.click()

    # Attendez que la page suivante se charge complètement (vous pouvez ajuster le temps d'attente si nécessaire)
    time.sleep(5)  # Exemple de temps d'attente de 5 secondes

    prenom_input_xpath = '//input[@id="field-id-9363320-0-firstname"]'
    prenom_input = driver.find_element(By.XPATH, prenom_input_xpath)
    prenom_input.send_keys(prenom_value)

    nom_input_xpath = '//input[@id="field-id-9363320-0-lastname"]'
    nom_input = driver.find_element(By.XPATH, nom_input_xpath)
    nom_input.send_keys(nom_value)

    texte_input_xpath = '//input[@name="9363320-0-14411222"]'
    texte_input = driver.find_element(By.XPATH, texte_input_xpath)
    texte_input.send_keys(texte_value)

    # Cliquez sur le bouton "Étape suivante"
    next_button = driver.find_element(By.XPATH, '//button[@data-ux="Forms_EVENT_NextStep"]')
    next_button.click()

    time.sleep(5)

    # Sélectionner l'élément du prénom par son ID
    prenom_input_id = 'field-firstname'
    prenom_input = driver.find_element(By.ID, prenom_input_id)
    prenom_input.send_keys(prenom_value)

    # Sélectionner l'élément de nom par son nouvel ID
    nom_input_id = 'field-lastname'
    nom_input = driver.find_element(By.ID, nom_input_id)
    nom_input.send_keys(nom_value)

    # Sélectionner l'élément d'e-mail en utilisant le chemin complet
    email_input = driver.find_element(By.CSS_SELECTOR, 'div[data-test="payer-personal-informations-email"] input[type="text"]')
    email_input.send_keys(email_value)

    # Cliquez sur le bouton "Étape suivante"
    next_button = driver.find_element(By.XPATH, '//button[@data-ux="Forms_EVENT_NextStep"]')
    next_button.click()

    time.sleep(5)

    # Utilisez JavaScript pour cocher une case à cocher par ID
    js_code = 'document.getElementById("field-cgu-consent").click();'
    driver.execute_script(js_code)

    # Cliquez sur le bouton "Valider"
    valider_button = driver.find_element(By.XPATH, '//button[contains(text(), "Valider")]')
    valider_button.click()

    time.sleep(10)

def main():
    url_page_avant = "url de votre site"
    driver = configure_browser()
    accept_cookies(driver, url_page_avant)

    workbook = load_workbook('C:/votre fichier.xlsx')
    sheet = workbook.active
    starting_row = 2

    for row in sheet.iter_rows(min_row=starting_row, values_only=True):
        quantity_value = row[3]
        email_value = row[2]
        nom_value = row[0]
        prenom_value = row[1]
        texte_value = row[4]

        fill_form(driver, quantity_value, prenom_value, nom_value, texte_value, email_value)

        # Après avoir soumis le formulaire avec succès
        print(f"Formulaire soumis avec succès pour {prenom_value} {nom_value}.")

        time.sleep(10)
        driver.get(url_page_avant)  # Revenir à la page d'accueil pour recommencer la boucle

    # Fermer le navigateur Selenium
    driver.quit()

